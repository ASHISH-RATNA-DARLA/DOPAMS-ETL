#!/usr/bin/env python3
"""
Person Deduplication Analysis Script - Database to Excel
Fetches person data from database, analyzes for duplicates
Uses ensemble fuzzy matching with multiple libraries
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import re
from typing import Dict, List, Tuple, Optional
import json
import psycopg2
from psycopg2.extras import RealDictCursor
import warnings

warnings.filterwarnings("ignore")

# Database connection
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgres://dopamasprd_ur:At0YM9pjD2rTmast4s@192.168.103.106:5432/dopamasuprddb",
)

# Try importing fuzzy matching libraries
try:
    from rapidfuzz import fuzz, distance

    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    print("âš ï¸  rapidfuzz not found. Install: pip install rapidfuzz")
    RAPIDFUZZ_AVAILABLE = False

try:
    from thefuzz import fuzz as thefuzz_fuzz
    from thefuzz import process as thefuzz_process

    THEFUZZ_AVAILABLE = True
except ImportError:
    print("âš ï¸  thefuzz not found. Install: pip install thefuzz python-Levenshtein")
    THEFUZZ_AVAILABLE = False

try:
    import dedupe

    DEDUPE_AVAILABLE = True
except ImportError:
    print("âš ï¸  dedupe not found. Install: pip install dedupe")
    DEDUPE_AVAILABLE = False

try:
    import textdistance

    TEXTDISTANCE_AVAILABLE = True
except ImportError:
    print("âš ï¸  textdistance not found. Install: pip install textdistance")
    TEXTDISTANCE_AVAILABLE = False

try:
    from difflib import SequenceMatcher

    DIFFLIB_AVAILABLE = True
except ImportError:
    DIFFLIB_AVAILABLE = False


# ==================== ADVANCED FUZZY MATCHER ====================
class AdvancedFuzzyMatcher:
    """Uses multiple fuzzy matching libraries for best results"""

    @staticmethod
    def normalize_name(name: str) -> str:
        """Normalize name for comparison"""
        if not name:
            return ""
        name = str(name).lower().strip()
        name = re.sub(r"\b(mr|mrs|ms|dr|md|s/o|d/o|w/o|shri|sri|prof)\b\.?", "", name)
        name = re.sub(r"[^\w\s]", "", name)
        name = re.sub(r"\s+", " ", name)
        return name.strip()

    @staticmethod
    def token_overlap_ratio(str1: str, str2: str) -> float:
        """Calculate token/word overlap ratio"""
        if not str1 or not str2:
            return 0.0

        tokens1 = set(AdvancedFuzzyMatcher.normalize_name(str1).split())
        tokens2 = set(AdvancedFuzzyMatcher.normalize_name(str2).split())

        if not tokens1 or not tokens2:
            return 0.0

        intersection = len(tokens1 & tokens2)
        union = len(tokens1 | tokens2)
        return intersection / union if union > 0 else 0.0

    @staticmethod
    def jaro_winkler_similarity(str1: str, str2: str) -> float:
        """Calculate Jaro-Winkler similarity"""
        if not str1 or not str2:
            return 0.0

        str1_norm = AdvancedFuzzyMatcher.normalize_name(str1)
        str2_norm = AdvancedFuzzyMatcher.normalize_name(str2)

        if str1_norm == str2_norm:
            return 1.0

        if RAPIDFUZZ_AVAILABLE:
            return fuzz.token_sort_ratio(str1_norm, str2_norm) / 100.0

        if THEFUZZ_AVAILABLE:
            return thefuzz_fuzz.token_sort_ratio(str1_norm, str2_norm) / 100.0

        if DIFFLIB_AVAILABLE:
            return SequenceMatcher(None, str1_norm, str2_norm).ratio()

        return 0.0

    @staticmethod
    def levenshtein_similarity(str1: str, str2: str) -> float:
        """Calculate Levenshtein distance based similarity"""
        if not str1 or not str2:
            return 0.0

        str1_norm = AdvancedFuzzyMatcher.normalize_name(str1)
        str2_norm = AdvancedFuzzyMatcher.normalize_name(str2)

        if str1_norm == str2_norm:
            return 1.0

        if RAPIDFUZZ_AVAILABLE:
            distance_val = distance.Levenshtein.distance(str1_norm, str2_norm)
            max_len = max(len(str1_norm), len(str2_norm))
            return 1.0 - (distance_val / max_len) if max_len > 0 else 0.0

        if THEFUZZ_AVAILABLE:
            return thefuzz_fuzz.ratio(str1_norm, str2_norm) / 100.0

        if TEXTDISTANCE_AVAILABLE:
            distance_val = textdistance.levenshtein(str1_norm, str2_norm)
            max_len = max(len(str1_norm), len(str2_norm))
            return 1.0 - (distance_val / max_len) if max_len > 0 else 0.0

        if DIFFLIB_AVAILABLE:
            return SequenceMatcher(None, str1_norm, str2_norm).ratio()

        return 0.0

    @staticmethod
    def jaccard_similarity(str1: str, str2: str) -> float:
        """Calculate Jaccard similarity"""
        if not str1 or not str2:
            return 0.0

        str1_norm = AdvancedFuzzyMatcher.normalize_name(str1)
        str2_norm = AdvancedFuzzyMatcher.normalize_name(str2)

        tokens1 = set(str1_norm.split())
        tokens2 = set(str2_norm.split())

        if not tokens1 or not tokens2:
            return 0.0

        if TEXTDISTANCE_AVAILABLE:
            return textdistance.jaccard(tokens1, tokens2)

        intersection = len(tokens1 & tokens2)
        union = len(tokens1 | tokens2)
        return intersection / union if union > 0 else 0.0

    @staticmethod
    def sorensen_similarity(str1: str, str2: str) -> float:
        """Calculate Sorensen-Dice similarity"""
        if not str1 or not str2:
            return 0.0

        str1_norm = AdvancedFuzzyMatcher.normalize_name(str1)
        str2_norm = AdvancedFuzzyMatcher.normalize_name(str2)

        if str1_norm == str2_norm:
            return 1.0

        if TEXTDISTANCE_AVAILABLE:
            return textdistance.sorensen(str1_norm, str2_norm)

        if RAPIDFUZZ_AVAILABLE:
            return fuzz.partial_ratio(str1_norm, str2_norm) / 100.0

        return AdvancedFuzzyMatcher.levenshtein_similarity(str1, str2)

    @staticmethod
    def ensemble_similarity(
        str1: str, str2: str, weights: Dict[str, float] = None
    ) -> Tuple[float, Dict]:
        """Ensemble method: combine multiple algorithms"""
        if weights is None:
            weights = {
                "levenshtein": 0.25,
                "jaro_winkler": 0.25,
                "token_overlap": 0.20,
                "jaccard": 0.15,
                "sorensen": 0.15,
            }

        breakdown = {}

        breakdown["levenshtein"] = AdvancedFuzzyMatcher.levenshtein_similarity(
            str1, str2
        )
        breakdown["jaro_winkler"] = AdvancedFuzzyMatcher.jaro_winkler_similarity(
            str1, str2
        )
        breakdown["token_overlap"] = AdvancedFuzzyMatcher.token_overlap_ratio(
            str1, str2
        )
        breakdown["jaccard"] = AdvancedFuzzyMatcher.jaccard_similarity(str1, str2)
        breakdown["sorensen"] = AdvancedFuzzyMatcher.sorensen_similarity(str1, str2)

        ensemble_score = sum(
            breakdown[method] * weights.get(method, 0) for method in breakdown.keys()
        )

        breakdown["ensemble"] = ensemble_score

        return ensemble_score, breakdown


# ==================== RELATION MATCHER ====================
class RelationMatcher:
    """Handles relation type matching"""

    RELATION_GROUPS = {
        "father": [
            "father",
            "s/o",
            "so",
            "bio father",
            "step father",
            "adopted father",
        ],
        "mother": [
            "mother",
            "d/o",
            "do",
            "bio mother",
            "step mother",
            "adopted mother",
        ],
        "spouse": ["spouse", "husband", "wife", "married", "partner"],
        "sibling": ["brother", "sister", "sibling", "bro", "sis"],
    }

    @staticmethod
    def normalize_relation(relation: str) -> str:
        """Normalize relation type"""
        if not relation:
            return ""
        rel = AdvancedFuzzyMatcher.normalize_name(relation)
        junk = ["unknown", "na", "n/a", "not mentioned", "not applicable", "-", ""]
        return "" if rel in junk else rel

    @staticmethod
    def match_relations(rel1: str, rel2: str) -> Tuple[float, Dict]:
        """Match relations"""
        rel1_norm = RelationMatcher.normalize_relation(rel1)
        rel2_norm = RelationMatcher.normalize_relation(rel2)

        flags = {}

        if not rel1_norm and not rel2_norm:
            flags["status"] = "both_missing"
            return 0.5, flags

        if not rel1_norm or not rel2_norm:
            flags["status"] = "one_missing"
            return 0.0, flags

        if rel1_norm == rel2_norm:
            flags["status"] = "exact_match"
            return 1.0, flags

        score, _ = AdvancedFuzzyMatcher.ensemble_similarity(rel1, rel2)
        flags["status"] = "fuzzy_match"

        return score, flags


# ==================== GENDER MATCHER ====================
class GenderMatcher:
    """Handles gender matching"""

    GENDER_VARIANTS = {
        "male": ["m", "male", "boy", "man"],
        "female": ["f", "female", "girl", "woman"],
    }

    @staticmethod
    def normalize_gender(gender: str) -> Optional[str]:
        """Normalize gender"""
        if not gender:
            return None
        g = AdvancedFuzzyMatcher.normalize_name(gender)
        for standard, variants in GenderMatcher.GENDER_VARIANTS.items():
            if g in variants:
                return standard
        return None

    @staticmethod
    def match_genders(g1: str, g2: str) -> Tuple[float, Dict]:
        """Match genders"""
        flags = {}

        norm_g1 = GenderMatcher.normalize_gender(g1)
        norm_g2 = GenderMatcher.normalize_gender(g2)

        if not norm_g1 and not norm_g2:
            flags["status"] = "both_missing"
            return 0.5, flags

        if not norm_g1 or not norm_g2:
            flags["status"] = "one_missing"
            return 0.3, flags

        if norm_g1 == norm_g2:
            flags["status"] = "match"
            return 1.0, flags

        flags["status"] = "mismatch"
        flags["conflict"] = f"{norm_g1} vs {norm_g2}"
        return 0.0, flags


# ==================== COMPREHENSIVE MATCHER ====================
class ComprehensiveMatcher:
    """Comprehensive person matching"""

    def __init__(self):
        self.fuzzy = AdvancedFuzzyMatcher()

    def match_persons(self, person1: Dict, person2: Dict) -> Tuple[float, Dict]:
        """Match persons using ensemble fuzzy matching"""
        breakdown = {}

        # 1. Full Name (50% weight)
        name1 = person1.get("full_name", "")
        name2 = person2.get("full_name", "")
        name_score, name_breakdown = self.fuzzy.ensemble_similarity(name1, name2)
        breakdown["full_name"] = {
            "score": round(name_score, 3),
            "weight": 0.50,
            "methods": name_breakdown,
        }

        # 2. Relative Name (30% weight)
        rel_name1 = person1.get("relative_name", "")
        rel_name2 = person2.get("relative_name", "")
        rel_score, rel_breakdown = self.fuzzy.ensemble_similarity(rel_name1, rel_name2)
        breakdown["relative_name"] = {
            "score": round(rel_score, 3),
            "weight": 0.30,
            "methods": rel_breakdown,
        }

        # 3. Gender (10% weight)
        gender_score, gender_flags = GenderMatcher.match_genders(
            person1.get("gender", ""), person2.get("gender", "")
        )
        breakdown["gender"] = {
            "score": round(gender_score, 3),
            "weight": 0.10,
            "flags": gender_flags,
        }

        # 4. Location (10% weight)
        locality1 = person1.get("present_locality_village", "")
        locality2 = person2.get("present_locality_village", "")
        loc_score, _ = self.fuzzy.ensemble_similarity(locality1, locality2)
        breakdown["location"] = {"score": round(loc_score, 3), "weight": 0.10}

        # Calculate overall weighted score
        overall = sum(
            v["score"] * v["weight"]
            for v in breakdown.values()
            if isinstance(v, dict) and "score" in v and "weight" in v
        )

        breakdown["overall"] = round(overall, 3)
        breakdown["match_threshold"] = 0.65
        breakdown["is_match"] = overall >= 0.65

        return overall, breakdown


# ==================== DATABASE ANALYZER ====================
class DatabaseAnalyzer:
    """Analyzes person data from database"""

    def __init__(self, db_url: str):
        self.db_url = db_url
        self.conn = None
        self.cursor = None
        self.df = None
        self.matcher = ComprehensiveMatcher()
        self.analysis_results = []

    def connect(self) -> bool:
        """Connect to database"""
        print(f"ðŸ”Œ Connecting to database...")
        try:
            self.conn = psycopg2.connect(self.db_url)
            self.cursor = self.conn.cursor(cursor_factory=RealDictCursor)
            print("âœ“ Connected successfully")
            return True
        except Exception as e:
            print(f"âœ— Error connecting to database: {e}")
            return False

    def fetch_persons_data(self) -> bool:
        """Fetch person data from database"""
        print(f"ðŸ“Š Fetching person data from database...")

        query = """
        SELECT 
            p.person_id,
            p.full_name,
            p.relative_name,
            p.gender,
            p.present_locality_village
        FROM persons p
        ORDER BY p.person_id
        """

        try:
            self.cursor.execute(query)
            records = self.cursor.fetchall()

            # Convert to dataframe
            self.df = pd.DataFrame(records)
            print(f"âœ“ Fetched {len(self.df)} person records")
            print(f"  Columns: {list(self.df.columns)}")
            return True
        except Exception as e:
            print(f"âœ— Error fetching data: {e}")
            return False

    def find_potential_duplicates(self) -> List[Dict]:
        """Find potential duplicate persons"""
        print(
            f"\nðŸ” Analyzing for potential duplicates using ensemble fuzzy matching..."
        )
        print(f"   Total person records: {len(self.df)}")

        matches = []
        analyzed = 0
        total_pairs = (len(self.df) * (len(self.df) - 1)) // 2

        for i in range(len(self.df)):
            for j in range(i + 1, len(self.df)):
                person1 = self.df.iloc[i].to_dict()
                person2 = self.df.iloc[j].to_dict()

                score, breakdown = self.matcher.match_persons(person1, person2)

                if score >= 0.50:
                    matches.append(
                        {
                            "idx1": i,
                            "idx2": j,
                            "person_id1": person1.get("person_id", "N/A"),
                            "person_id2": person2.get("person_id", "N/A"),
                            "person1_name": person1.get("full_name", "N/A"),
                            "person2_name": person2.get("full_name", "N/A"),
                            "score": score,
                            "breakdown": breakdown,
                            "person1": person1,
                            "person2": person2,
                        }
                    )

                analyzed += 1
                if analyzed % 10000 == 0:
                    pct = (analyzed / total_pairs) * 100
                    print(f"   Analyzed {analyzed}/{total_pairs} pairs ({pct:.1f}%)...")

        matches.sort(key=lambda x: x["score"], reverse=True)
        print(f"\nâœ“ Found {len(matches)} potential matches (score â‰¥ 0.50)")
        return matches

    def create_output_workbook(self, matches: List[Dict]):
        """Create comprehensive Excel workbook"""
        print(f"\nðŸ“Š Creating Excel workbook with analysis...")

        output_file = f"person_deduplication_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Sheet 1: Executive Summary & Documentation
            self._create_summary_sheet(writer)

            # Sheet 2: Original Data
            self.df.to_excel(writer, sheet_name="Original Data", index=False)

            # Sheet 3: Duplicate Matches
            if matches:
                matches_data = []
                for match in matches:
                    bd = match["breakdown"]
                    matches_data.append(
                        {
                            "Person ID 1": match["person_id1"],
                            "Person ID 2": match["person_id2"],
                            "Name 1": match["person1_name"],
                            "Name 2": match["person2_name"],
                            "Match Score %": f"{match['score']*100:.1f}%",
                            "Full Name %": f"{bd['full_name']['score']*100:.1f}%",
                            "Relative Name %": f"{bd['relative_name']['score']*100:.1f}%",
                            "Gender %": f"{bd['gender']['score']*100:.1f}%",
                            "Location %": f"{bd['location']['score']*100:.1f}%",
                            "Verdict": (
                                "STRONG DUPLICATE"
                                if match["score"] >= 0.80
                                else (
                                    "LIKELY DUPLICATE"
                                    if match["score"] >= 0.65
                                    else "POSSIBLE MATCH"
                                )
                            ),
                        }
                    )

                matches_df = pd.DataFrame(matches_data)
                matches_df.to_excel(writer, sheet_name="Duplicate Matches", index=False)
            else:
                pd.DataFrame({"Result": ["No duplicates found"]}).to_excel(
                    writer, sheet_name="Duplicate Matches", index=False
                )

            # Sheet 4: Field Quality Analysis
            field_analysis = self._generate_field_analysis()
            field_analysis.to_excel(writer, sheet_name="Field Quality", index=False)

            # Sheet 5: Matching Methodology
            self._create_methodology_sheet(writer)

            # Sheet 6: Statistics
            stats = self._generate_statistics(matches)
            stats.to_excel(writer, sheet_name="Statistics", index=False)

        self._format_workbook(output_file)

        print(f"âœ“ Saved to: {output_file}")
        return output_file

    def _generate_field_analysis(self) -> pd.DataFrame:
        """Analyze field quality"""
        analysis = []
        for col in self.df.columns:
            filled = self.df[col].notna().sum()
            empty = len(self.df) - filled
            completeness = (filled / len(self.df)) * 100 if len(self.df) > 0 else 0
            analysis.append(
                {
                    "Field": col,
                    "Total Records": len(self.df),
                    "Filled": filled,
                    "Empty": empty,
                    "Completeness %": f"{completeness:.1f}%",
                }
            )
        return pd.DataFrame(analysis)

    def _generate_statistics(self, matches: List[Dict]) -> pd.DataFrame:
        """Generate statistics"""
        strong = len([m for m in matches if m["score"] >= 0.80])
        likely = len([m for m in matches if 0.65 <= m["score"] < 0.80])
        possible = len([m for m in matches if 0.50 <= m["score"] < 0.65])

        return pd.DataFrame(
            [
                {"Metric": "Total Person Records", "Value": len(self.df)},
                {
                    "Metric": "Total Possible Pairs",
                    "Value": (len(self.df) * (len(self.df) - 1)) // 2,
                },
                {"Metric": "Strong Matches (â‰¥0.80)", "Value": strong},
                {"Metric": "Likely Matches (0.65-0.80)", "Value": likely},
                {"Metric": "Possible Matches (0.50-0.65)", "Value": possible},
                {"Metric": "Total Matches Found", "Value": len(matches)},
                {
                    "Metric": "Analysis Date",
                    "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                },
            ]
        )

    def _create_summary_sheet(self, writer):
        """Create executive summary sheet with documentation"""
        summary_text = [
            ["TOYSTACK DOPAMS DATA ANALYSIS"],
            ["Person Deduplication Analysis Report"],
            [""],
            ["ANALYSIS OVERVIEW"],
            [
                "This report analyzes the persons database for duplicate and similar records."
            ],
            [
                "Using advanced fuzzy matching algorithms to identify potential duplicate persons"
            ],
            [
                "across the entire database, helping maintain data quality and integrity."
            ],
            [""],
            ["REPORT GENERATED"],
            [f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            [f"Total Records Analyzed: {len(self.df)}"],
            [""],
            ["DUPLICATE CLASSIFICATION"],
            ["STRONG DUPLICATE (â‰¥0.80): High confidence matches - likely same person"],
            [
                "LIKELY DUPLICATE (0.65-0.80): Good confidence matches - probably same person"
            ],
            [
                "POSSIBLE MATCH (0.50-0.65): Low confidence matches - may need manual review"
            ],
            [""],
            ["MATCHING METHODOLOGY"],
            ["The system uses ensemble fuzzy matching combining 5 algorithms:"],
            ["  â€¢ Levenshtein Distance (25%) - Character-level string similarity"],
            ["  â€¢ Jaro-Winkler (25%) - Phonetic similarity for names"],
            ["  â€¢ Token Overlap (20%) - Word-level matching"],
            ["  â€¢ Jaccard Similarity (15%) - Set-based word comparison"],
            ["  â€¢ Sorensen-Dice (15%) - Gestalt pattern matching"],
            [""],
            ["FIELD WEIGHTS IN MATCHING"],
            ["  â€¢ Full Name: 50% (HIGHEST PRIORITY)"],
            ["  â€¢ Relative Name (Parent): 30% (VERY HIGH PRIORITY)"],
            ["  â€¢ Gender: 10% (MEDIUM PRIORITY)"],
            ["  â€¢ Location: 10% (Supporting)"],
            [""],
            ["KEY FEATURES"],
            [
                "âœ“ Automatic name normalization (removes prefixes like S/O, D/O, Dr., Mr.)"
            ],
            ["âœ“ Handles typos and spelling variations"],
            ["âœ“ Space and punctuation normalization"],
            ["âœ“ Gender validation and cross-checking"],
            ["âœ“ Geographic location matching"],
            [""],
            ["SHEET DESCRIPTIONS"],
            ["Sheet 1 (Executive Summary): This overview and documentation"],
            [
                "Sheet 2 (Original Data): Complete list of all person records from database"
            ],
            ["Sheet 3 (Duplicate Matches): Ranked list of potential duplicate persons"],
            ["Sheet 4 (Field Quality): Data completeness analysis per field"],
            ["Sheet 5 (Methodology): Detailed explanation of matching algorithms"],
            ["Sheet 6 (Statistics): Summary statistics and metrics"],
            [""],
            ["RECOMMENDATIONS"],
            ["1. Review STRONG DUPLICATE matches first (highest confidence)"],
            ["2. Manually verify LIKELY DUPLICATE matches before action"],
            ["3. Consider POSSIBLE MATCH as reference only"],
            ["4. Update missing fields (gender, location) to improve matching"],
            ["5. Standardize name formats for better results"],
            [""],
            ["QUALITY METRICS"],
            ["â€¢ Match Confidence Threshold: 65% (0.65)"],
            ["â€¢ Name Similarity Threshold: 80%"],
            ["â€¢ Typo Tolerance: Up to 3 character differences"],
            ["â€¢ Gender Matching: Exact match or both missing"],
            ["â€¢ Location Matching: Fuzzy matching with ensemble algorithms"],
        ]

        summary_df = pd.DataFrame(summary_text)
        summary_df.to_excel(
            writer, sheet_name="Executive Summary", header=False, index=False
        )

    def _create_methodology_sheet(self, writer):
        """Create detailed methodology sheet"""
        methodology_text = [
            ["MATCHING ALGORITHMS EXPLAINED"],
            [""],
            ["1. LEVENSHTEIN DISTANCE (25% weight)"],
            [
                "   Definition: Minimum number of single-character edits (insertions, deletions, substitutions)"
            ],
            [
                "   Example: 'Rajesh' vs 'Rajish' = 1 edit (substitution eâ†’i) = 85% match"
            ],
            ["   Use Case: Catches typos and spelling variations in names"],
            [""],
            ["2. JARO-WINKLER SIMILARITY (25% weight)"],
            [
                "   Definition: Variant of Jaro distance that gives more weight to matching strings from beginning"
            ],
            ["   Example: 'Abdul' vs 'Abdel' = 89% match (phonetically similar)"],
            ["   Use Case: Excellent for catching phonetic similarities in names"],
            [""],
            ["3. TOKEN OVERLAP RATIO (20% weight)"],
            ["   Definition: Percentage of common words between two names"],
            [
                "   Example: 'Raj Kumar Singh' vs 'Singh Raj Kumar' = 100% (all words present)"
            ],
            ["   Use Case: Handles name ordering variations"],
            [""],
            ["4. JACCARD SIMILARITY (15% weight)"],
            ["   Definition: Intersection divided by union of character sets"],
            [
                "   Example: 'Mohammad' vs 'Mohammed' = 75% match (similar character set)"
            ],
            ["   Use Case: Useful for accent and spelling variations"],
            [""],
            ["5. SORENSEN-DICE COEFFICIENT (15% weight)"],
            ["   Definition: Similar to Jaccard but weights overlap more heavily"],
            ["   Example: Detects partial matches with word segments"],
            ["   Use Case: Good at finding closely related strings"],
            [""],
            ["ENSEMBLE APPROACH"],
            [
                "The system calculates all 5 algorithms and combines them using weighted average:"
            ],
            [
                "Final Score = (LevÃ—0.25) + (JWÃ—0.25) + (TokenÃ—0.20) + (JaccardÃ—0.15) + (SorensenÃ—0.15)"
            ],
            [""],
            ["FIELD-LEVEL MATCHING"],
            [""],
            ["FULL NAME (50% Weight) - HIGHEST PRIORITY"],
            ["â€¢ Uses all 5 ensemble methods"],
            ["â€¢ Normalizes: removes prefixes (Mr., Dr., S/O), special characters"],
            ["â€¢ Case-insensitive matching"],
            ["â€¢ Handles abbreviations"],
            [""],
            ["RELATIVE NAME (Parent Name) (30% Weight) - VERY HIGH"],
            ["â€¢ Uses all 5 ensemble methods"],
            ["â€¢ Strips relation prefixes (S/O = Son Of, D/O = Daughter Of)"],
            ["â€¢ Handles name abbreviations (Abdul R. vs Abdul Rahman)"],
            ["â€¢ Semantic matching of name variants"],
            [""],
            ["GENDER (10% Weight) - MEDIUM"],
            ["â€¢ Binary matching: Male vs Female"],
            ["â€¢ Variant handling: M/Male/Boy/Man all treated as Male"],
            ["â€¢ F/Female/Girl/Woman all treated as Female"],
            ["â€¢ Missing value handling: 0.5 score if both missing"],
            [""],
            ["LOCATION (10% Weight)"],
            ["â€¢ Uses ensemble fuzzy matching on locality/village names"],
            ["â€¢ Handles spelling variations in place names"],
            ["â€¢ Provides geographic context for matching"],
            [""],
            ["CONFIDENCE THRESHOLDS"],
            ["Score â‰¥ 0.80 = STRONG DUPLICATE (95% confidence)"],
            ["Score 0.65-0.80 = LIKELY DUPLICATE (75% confidence)"],
            ["Score 0.50-0.65 = POSSIBLE MATCH (50% confidence)"],
            ["Score < 0.50 = Not reported (too uncertain)"],
        ]

        methodology_df = pd.DataFrame(methodology_text)
        methodology_df.to_excel(
            writer, sheet_name="Methodology", header=False, index=False
        )

    def _format_workbook(self, output_file: str):
        """Apply formatting to workbook"""
        try:
            wb = openpyxl.load_workbook(output_file)

            # Format Executive Summary sheet
            if "Executive Summary" in wb.sheetnames:
                ws = wb["Executive Summary"]
                title_fill = PatternFill(
                    start_color="1F4E78", end_color="1F4E78", fill_type="solid"
                )
                title_font = Font(bold=True, color="FFFFFF", size=14)
                subtitle_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                subtitle_font = Font(bold=True, color="FFFFFF", size=11)

                ws.column_dimensions["A"].width = 70

                for idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=ws.max_row), 1
                ):
                    cell = row[0]
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

                    # Title
                    if idx == 1:
                        cell.fill = title_fill
                        cell.font = title_font
                        ws.row_dimensions[idx].height = 25
                    # Subtitles and section headers
                    elif any(
                        x in str(cell.value)
                        for x in [
                            "ANALYSIS OVERVIEW",
                            "REPORT GENERATED",
                            "DUPLICATE CLASSIFICATION",
                            "MATCHING METHODOLOGY",
                            "FIELD WEIGHTS",
                            "KEY FEATURES",
                            "SHEET DESCRIPTIONS",
                            "RECOMMENDATIONS",
                            "QUALITY METRICS",
                        ]
                    ):
                        cell.fill = subtitle_fill
                        cell.font = subtitle_font
                        ws.row_dimensions[idx].height = 18

            # Format Methodology sheet
            if "Methodology" in wb.sheetnames:
                ws = wb["Methodology"]
                ws.column_dimensions["A"].width = 85

                for idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=ws.max_row), 1
                ):
                    cell = row[0]
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

                    # Headers
                    if "EXPLAINED" in str(cell.value) or any(
                        x in str(cell.value)
                        for x in [
                            "LEVENSHTEIN",
                            "JARO-WINKLER",
                            "TOKEN OVERLAP",
                            "JACCARD",
                            "SORENSEN",
                            "ENSEMBLE",
                            "FIELD-LEVEL",
                            "CONFIDENCE",
                        ]
                    ):
                        cell.fill = subtitle_fill
                        cell.font = subtitle_font
                        ws.row_dimensions[idx].height = 16

            # Format Duplicate Matches sheet
            if "Duplicate Matches" in wb.sheetnames:
                ws = wb["Duplicate Matches"]

                header_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                # Data formatting
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=2, max_row=ws.max_row), 2
                ):
                    for col_idx, cell in enumerate(row, 1):
                        cell.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=True
                        )

                        # Color code by verdict
                        if "Verdict" in ws[1][col_idx - 1].value or col_idx == len(row):
                            if "STRONG" in str(cell.value):
                                cell.fill = PatternFill(
                                    start_color="C6EFCE",
                                    end_color="C6EFCE",
                                    fill_type="solid",
                                )
                                cell.font = Font(color="006100", bold=True)
                            elif "LIKELY" in str(cell.value):
                                cell.fill = PatternFill(
                                    start_color="FFF2CC",
                                    end_color="FFF2CC",
                                    fill_type="solid",
                                )
                                cell.font = Font(color="9C6500", bold=True)
                            else:
                                cell.fill = PatternFill(
                                    start_color="FCE4D6",
                                    end_color="FCE4D6",
                                    fill_type="solid",
                                )
                                cell.font = Font(color="C65911", bold=True)

                # Adjust column widths
                ws.column_dimensions["A"].width = 15
                ws.column_dimensions["B"].width = 15
                ws.column_dimensions["C"].width = 20
                ws.column_dimensions["D"].width = 20
                ws.column_dimensions["E"].width = 12

            wb.save(output_file)
        except Exception as e:
            print(f"âš ï¸  Warning: Could not format workbook: {e}")

    def disconnect(self):
        """Close database connection"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
        print("âœ“ Disconnected from database")

    def run(self):
        """Main execution"""
        try:
            if not self.connect():
                return

            if not self.fetch_persons_data():
                return

            matches = self.find_potential_duplicates()
            self.create_output_workbook(matches)

            print("\n" + "=" * 70)
            print("âœ… Analysis Complete!")
            print("=" * 70)
            print(f"\nFuzzy Matching Methods Used:")
            print(f"  â€¢ Levenshtein Distance (25% weight)")
            print(f"  â€¢ Jaro-Winkler (25% weight)")
            print(f"  â€¢ Token Overlap (20% weight)")
            print(f"  â€¢ Jaccard Similarity (15% weight)")
            print(f"  â€¢ Sorensen-Dice (15% weight)")
            print(f"\nLibraries Active:")
            print(f"  âœ“ rapidfuzz" if RAPIDFUZZ_AVAILABLE else f"  âœ— rapidfuzz")
            print(f"  âœ“ thefuzz" if THEFUZZ_AVAILABLE else f"  âœ— thefuzz")
            print(f"  âœ“ dedupe" if DEDUPE_AVAILABLE else f"  âœ— dedupe")
            print(
                f"  âœ“ textdistance" if TEXTDISTANCE_AVAILABLE else f"  âœ— textdistance"
            )
            print("\n")

        except Exception as e:
            print(f"\nâŒ Error: {e}")
            import traceback

            traceback.print_exc()
        finally:
            self.disconnect()


# ==================== MAIN ====================
if __name__ == "__main__":
    print("=" * 70)
    print("Person Deduplication Analysis - Database to Excel")
    print("Using Ensemble Fuzzy Matching with Multiple Libraries")
    print("=" * 70 + "\n")

    db_url = os.getenv("DATABASE_URL", DATABASE_URL)

    analyzer = DatabaseAnalyzer(db_url)
    analyzer.run()

