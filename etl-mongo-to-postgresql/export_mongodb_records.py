"""
Export Script for MongoDB-Migrated Records
Exports 50 records from each PostgreSQL table that were migrated from MongoDB
"""

import os
import sys
import logging
import csv
import shutil
from datetime import datetime
from typing import List, Dict, Any, Set, Tuple

import pymongo
from pymongo import MongoClient
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'export_mongodb_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class MongoDBRecordsExporter:
    """Export MongoDB-migrated records from PostgreSQL"""
    
    def __init__(self):
        """Initialize connections"""
        self.conn = None
        self.mongo_client = None
        self.mongo_db = None
        self.mongodb_crime_ids: Set[str] = set()
        self.output_folder = 'review-old-records'
    
    def setup_output_folder(self):
        """Create output folder, delete if exists"""
        try:
            # Delete folder if it exists
            if os.path.exists(self.output_folder):
                logger.info(f"Deleting existing folder: {self.output_folder}")
                shutil.rmtree(self.output_folder)
            
            # Create new folder
            os.makedirs(self.output_folder, exist_ok=True)
            logger.info(f"Created output folder: {self.output_folder}")
            return True
        except Exception as e:
            logger.error(f"Error setting up output folder: {e}")
            return False
        
    def connect_mongodb(self):
        """Connect to MongoDB and get actual MongoDB ObjectIds"""
        try:
            mongo_uri = os.getenv('MONGO_URI')
            mongo_db_name = os.getenv('MONGO_DB_NAME')
            mongo_collection_name = os.getenv('MONGO_COLLECTION_NAME')
            
            if not mongo_uri or not mongo_db_name:
                raise ValueError("MONGO_URI and MONGO_DB_NAME must be set in .env")
                
            self.mongo_client = MongoClient(mongo_uri)
            self.mongo_db = self.mongo_client[mongo_db_name]
            
            # Test connection
            self.mongo_client.admin.command('ping')
            logger.info(f"Connected to MongoDB: {mongo_db_name}")
            
            # Get all MongoDB ObjectIds from the collection
            collection = self.mongo_db[mongo_collection_name]
            logger.info(f"Fetching MongoDB ObjectIds from collection: {mongo_collection_name}")
            
            # Get first 50 records' ObjectIds (as strings)
            cursor = collection.find({}, {'_id': 1}).limit(50)
            for doc in cursor:
                mongo_id = doc.get('_id')
                if mongo_id:
                    # Convert ObjectId to string
                    crime_id = str(mongo_id)
                    self.mongodb_crime_ids.add(crime_id)
            
            logger.info(f"Found {len(self.mongodb_crime_ids)} MongoDB ObjectIds to match")
            return True
        except Exception as e:
            logger.error(f"Failed to connect to MongoDB: {e}")
            return False
    
    def connect_postgresql(self):
        """Connect to PostgreSQL"""
        try:
            pg_config = {
                'host': os.getenv('POSTGRES_HOST'),
                'port': os.getenv('POSTGRES_PORT'),
                'database': os.getenv('POSTGRES_DB'),
                'user': os.getenv('POSTGRES_USER'),
                'password': os.getenv('POSTGRES_PASSWORD')
            }
            
            if not all([pg_config['host'], pg_config['database'], pg_config['user'], pg_config['password']]):
                raise ValueError("PostgreSQL connection details must be set in .env")
                
            self.conn = psycopg2.connect(**pg_config)
            logger.info(f"Connected to PostgreSQL: {pg_config['database']}")
            return True
        except Exception as e:
            logger.error(f"Failed to connect to PostgreSQL: {e}")
            return False
            
    def get_table_data(self, table_name: str, query: str, params: tuple = None) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get table data and return column names and rows"""
        cursor = self.conn.cursor(cursor_factory=RealDictCursor)
        try:
            cursor.execute(query, params or ())
            rows = cursor.fetchall()
            
            if not rows:
                logger.warning(f"No records found for table: {table_name}")
                return [], []
            
            # Get column names
            column_names = [desc[0] for desc in cursor.description]
            
            # Convert rows to list of dicts and handle None values
            data_rows = []
            for row in rows:
                row_dict = dict(row)
                # Convert None to empty string and datetime to string
                for key, value in row_dict.items():
                    if value is None:
                        row_dict[key] = ''
                    elif isinstance(value, datetime):
                        row_dict[key] = value.isoformat()
                    else:
                        # Convert other types to string for Excel
                        row_dict[key] = str(value) if value is not None else ''
                data_rows.append(row_dict)
            
            logger.info(f"Retrieved {len(data_rows)} records from {table_name}")
            return column_names, data_rows
            
        except Exception as e:
            logger.error(f"Error getting data from {table_name}: {e}")
            return [], []
        finally:
            cursor.close()
    
    def export_to_excel(self, data_dict: Dict[str, Tuple[List[str], List[Dict[str, Any]]]], filename: str = None):
        """Export all table data to a single Excel file with multiple sheets"""
        if not filename:
            filename = f'mongodb_records_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Ensure filename is in the output folder
        filename = os.path.join(self.output_folder, filename)
        
        try:
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create a sheet for each table
            for table_name, (column_names, rows) in data_dict.items():
                if not column_names or not rows:
                    logger.warning(f"Skipping {table_name} - no data")
                    continue
                
                # Create sheet (Excel sheet names are limited to 31 characters and cannot contain: / \ ? * [ ] :)
                # Replace invalid characters with underscore
                invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
                sheet_name = table_name
                for char in invalid_chars:
                    sheet_name = sheet_name.replace(char, '_')
                # Truncate to 31 characters if needed
                sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                ws = wb.create_sheet(title=sheet_name)
                
                # Write headers
                for col_idx, col_name in enumerate(column_names, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = col_name
                    cell.font = cell.font.copy(bold=True)
                
                # Write data rows
                for row_idx, row_dict in enumerate(rows, start=2):
                    for col_idx, col_name in enumerate(column_names, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = row_dict.get(col_name, '')
                        cell.value = value
                
                # Auto-adjust column widths
                for col_idx, col_name in enumerate(column_names, start=1):
                    max_length = len(str(col_name))
                    for row in ws.iter_rows(min_row=2, max_row=min(len(rows) + 1, 100), min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    # Set column width (with a max of 50)
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                
                logger.info(f"Added sheet '{sheet_name}' with {len(rows)} rows")
            
            # Save workbook
            wb.save(filename)
            logger.info(f"Excel file created: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}", exc_info=True)
            return False
    
    def get_mongodb_crime_ids_tuple(self) -> tuple:
        """Get tuple of MongoDB crime IDs for SQL IN clause"""
        if not self.mongodb_crime_ids:
            return tuple()
        return tuple(self.mongodb_crime_ids)
    
    def get_hierarchy_data(self, limit: int = 50) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get hierarchy records linked to MongoDB crimes"""
        if not self.mongodb_crime_ids:
            logger.warning("No MongoDB crime IDs found, skipping hierarchy")
            return [], []
        
        ids_tuple = self.get_mongodb_crime_ids_tuple()
        # Create placeholders for IN clause
        placeholders = ','.join(['%s'] * len(ids_tuple))
        query = f"""
            SELECT DISTINCT h.*
            FROM hierarchy h
            INNER JOIN crimes c ON h.ps_code = c.ps_code
            WHERE c.crime_id IN ({placeholders})
            LIMIT %s
        """
        params = ids_tuple + (limit,)
        return self.get_table_data('hierarchy', query, params)
    
    def get_crimes_data(self, limit: int = 50) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get crimes records that are MongoDB-migrated"""
        if not self.mongodb_crime_ids:
            logger.warning("No MongoDB crime IDs found, skipping crimes")
            return [], []
        
        ids_tuple = self.get_mongodb_crime_ids_tuple()
        # Create placeholders for IN clause
        placeholders = ','.join(['%s'] * len(ids_tuple))
        query = f"""
            SELECT *
            FROM crimes
            WHERE crime_id IN ({placeholders})
            LIMIT %s
        """
        params = ids_tuple + (limit,)
        return self.get_table_data('crimes', query, params)
    
    def get_accused_data(self, limit: int = 50) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get accused records linked to MongoDB crimes"""
        if not self.mongodb_crime_ids:
            logger.warning("No MongoDB crime IDs found, skipping accused")
            return [], []
        
        ids_tuple = self.get_mongodb_crime_ids_tuple()
        # Create placeholders for IN clause
        placeholders = ','.join(['%s'] * len(ids_tuple))
        query = f"""
            SELECT a.*
            FROM accused a
            INNER JOIN crimes c ON a.crime_id = c.crime_id
            WHERE c.crime_id IN ({placeholders})
            LIMIT %s
        """
        params = ids_tuple + (limit,)
        return self.get_table_data('accused', query, params)
    
    def get_persons_data(self, limit: int = 50) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get persons records linked to accused from MongoDB crimes"""
        if not self.mongodb_crime_ids:
            logger.warning("No MongoDB crime IDs found, skipping persons")
            return [], []
        
        ids_tuple = self.get_mongodb_crime_ids_tuple()
        # Create placeholders for IN clause
        placeholders = ','.join(['%s'] * len(ids_tuple))
        query = f"""
            SELECT DISTINCT p.*
            FROM persons p
            INNER JOIN accused a ON p.person_id = a.person_id
            INNER JOIN crimes c ON a.crime_id = c.crime_id
            WHERE c.crime_id IN ({placeholders})
            LIMIT %s
        """
        params = ids_tuple + (limit,)
        return self.get_table_data('persons', query, params)
    
    def get_brief_facts_drugs_data(self, limit: int = 50) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get brief_facts_drugs records linked to MongoDB crimes"""
        if not self.mongodb_crime_ids:
            logger.warning("No MongoDB crime IDs found, skipping brief_facts_drugs")
            return [], []
        
        ids_tuple = self.get_mongodb_crime_ids_tuple()
        # Create placeholders for IN clause
        placeholders = ','.join(['%s'] * len(ids_tuple))
        query = f"""
            SELECT bfd.*
            FROM brief_facts_drugs bfd
            INNER JOIN crimes c ON bfd.crime_id = c.crime_id
            WHERE c.crime_id IN ({placeholders})
            LIMIT %s
        """
        params = ids_tuple + (limit,)
        return self.get_table_data('brief_facts_drugs', query, params)
    
    def get_interrogation_report_data(self, limit: int = 50) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Get old_interragation_report records linked to MongoDB crimes"""
        if not self.mongodb_crime_ids:
            logger.warning("No MongoDB crime IDs found, skipping interrogation_report")
            return [], []
        
        ids_tuple = self.get_mongodb_crime_ids_tuple()
        # Create placeholders for IN clause
        placeholders = ','.join(['%s'] * len(ids_tuple))
        query = f"""
            SELECT oir.*
            FROM old_interragation_report oir
            INNER JOIN crimes c ON oir.crime_id = c.crime_id
            WHERE c.crime_id IN ({placeholders})
            LIMIT %s
        """
        params = ids_tuple + (limit,)
        return self.get_table_data('old_interragation_report', query, params)
    
    def export_all_tables(self, limit: int = 50):
        """Export 50 records from all tables to a single Excel file"""
        logger.info(f"Starting export of {limit} records from each table...")
        
        # Collect data from all tables
        data_dict = {}
        
        logger.info("Collecting data from hierarchy...")
        columns, rows = self.get_hierarchy_data(limit)
        if columns and rows:
            data_dict['hierarchy'] = (columns, rows)
        
        logger.info("Collecting data from crimes...")
        columns, rows = self.get_crimes_data(limit)
        if columns and rows:
            data_dict['crimes'] = (columns, rows)
        
        logger.info("Collecting data from accused...")
        columns, rows = self.get_accused_data(limit)
        if columns and rows:
            data_dict['accused'] = (columns, rows)
        
        logger.info("Collecting data from persons...")
        columns, rows = self.get_persons_data(limit)
        if columns and rows:
            data_dict['persons'] = (columns, rows)
        
        logger.info("Collecting data from brief_facts_drugs...")
        columns, rows = self.get_brief_facts_drugs_data(limit)
        if columns and rows:
            data_dict['brief_facts_drugs'] = (columns, rows)
        
        logger.info("Collecting data from old_interragation_report...")
        columns, rows = self.get_interrogation_report_data(limit)
        if columns and rows:
            data_dict['old_interragation_report'] = (columns, rows)
        
        # Export to single Excel file
        if data_dict:
            logger.info(f"\nCreating Excel file with {len(data_dict)} sheets...")
            success = self.export_to_excel(data_dict)
            
            logger.info("\n" + "="*60)
            logger.info("Export Summary:")
            logger.info("="*60)
            for table_name, (columns, rows) in data_dict.items():
                logger.info(f"{table_name:30s} ✓ {len(rows)} records")
            logger.info("="*60)
            
            return success
        else:
            logger.warning("No data collected from any table")
            return False
    
    def close(self):
        """Close database connections"""
        if self.conn:
            self.conn.close()
            logger.info("PostgreSQL connection closed")
        if self.mongo_client:
            self.mongo_client.close()
            logger.info("MongoDB connection closed")


def main():
    """Main function"""
    exporter = MongoDBRecordsExporter()
    
    try:
        # Setup output folder (delete if exists, create new)
        if not exporter.setup_output_folder():
            logger.error("Failed to setup output folder. Exiting.")
            return 1
        
        # Connect to MongoDB first to get actual MongoDB ObjectIds
        if not exporter.connect_mongodb():
            logger.error("Failed to connect to MongoDB. Exiting.")
            return 1
        
        # Connect to PostgreSQL
        if not exporter.connect_postgresql():
            logger.error("Failed to connect to PostgreSQL. Exiting.")
            return 1
        
        # Export all tables (only records matching MongoDB ObjectIds)
        exporter.export_all_tables(limit=50)
        
        logger.info(f"\nExport completed! Check the Excel file in the '{exporter.output_folder}' folder.")
        logger.info(f"Exported records matching {len(exporter.mongodb_crime_ids)} MongoDB ObjectIds")
        return 0
        
    except Exception as e:
        logger.error(f"Error during export: {e}", exc_info=True)
        return 1
    finally:
        exporter.close()


if __name__ == '__main__':
    sys.exit(main())


